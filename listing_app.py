import streamlit as st
import pandas as pd
import io, base64, json, re, openpyxl
from openai import OpenAI
from openpyxl.styles import Font, Alignment

# --- 1. 核心工具函數 (物理鎖定規則) ---
def strict_clean(text):
    if not text: return ""
    # 移除 JSON 符號及 AI 佔位詞，解決圖 d7cb 的逗號和佔位符問題
    text = re.sub(r"[\[\]'\"']", "", str(text))
    blacklist = {'word1', 'word2', 'fake', 'placeholder', 'detailed', 'rich', 'title'}
    words = text.split()
    return " ".join([w for w in words if w.lower() not in blacklist]).strip()

def format_kw_strict(raw_text):
    """關鍵詞規則：僅空格分隔，限長 245"""
    clean = re.sub(r'[^a-zA-Z0-9\s]', ' ', str(raw_text).lower())
    seen, res = set(), []
    for w in clean.split():
        if w not in seen and len(w) > 1:
            res.append(w)
            seen.add(w)
    return " ".join(res)[:245]

# --- 2. 頁面配置 ---
st.set_page_config(page_title="亞馬遜 AI 全能系統 V14.5", layout="wide")
api_key = st.secrets.get("OPENAI_API_KEY") or ""

# --- 3. 功能導航 (側邊欄) ---
st.sidebar.title("🚀 功能導航")
# 使用 radio 進行物理隔離，避免代碼衝突
mode = st.sidebar.radio("請選擇操作模式：", ["批量上架 (圖片分析)", "站點搬運 (US ➔ UK)"])

# ==========================================
# 模式一：批量上架 (保持 Row 4 鎖定邏輯)
# ==========================================
if mode == "批量上架 (圖片分析)":
    st.header("🎨 AI 視覺分析上架模塊")
    
    with st.sidebar:
        st.subheader("⚙️ 規格鎖定")
        brand = st.text_input("品牌名稱", value="AMAZING WALL", key="v145_brand")
        s1, p1, n1 = st.text_input("尺寸 1", "16x24\"", key="s1"), st.text_input("價格 1", "12.99", key="p1"), "001"
        s2, p2, n2 = st.text_input("尺寸 2", "24x36\"", key="s2"), st.text_input("價格 2", "16.99", key="p2"), "002"
        s3, p3, n3 = st.text_input("尺寸 3", "32x48\"", key="s3"), st.text_input("價格 3", "19.99", key="p3"), "003"

    if 'v145_rows' not in st.session_state: st.session_state.v145_rows = 1
    sku_inputs = []
    for i in range(st.session_state.v145_rows):
        with st.expander(f"款式 {i+1}", expanded=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                pfx = st.text_input("SKU 前綴", key=f"v145_pfx_{i}")
                img = st.file_uploader("分析圖", key=f"v145_img_{i}")
            with c2:
                mu = st.text_input("主圖 URL", key=f"v145_mu_{i}")
                ou = st.text_area("附圖集", key=f"v145_ou_{i}")
            with c3:
                u1 = st.text_input(f"{s1} 圖", key=f"v145_u1_{i}")
                u2 = st.text_input(f"{s2} 圖", key=f"v145_u2_{i}")
                u3 = st.text_input(f"{s3} 圖", key=f"v145_u3_{i}")
            sku_inputs.append({"pfx": pfx, "img": img, "main": mu, "sz_urls": [u1, u2, u3]})

    if st.button("➕ 增加款式"):
        st.session_state.v145_rows += 1
        st.rerun()

    user_kw = st.text_area("Search Terms 詞庫")
    uploaded_tpl = st.file_uploader("📂 上傳 Amazon 模板", type=['xlsx', 'xlsm'], key="v145_tpl")

    if st.button("🚀 啟動 AI 填充", type="primary"):
        if not uploaded_tpl or not api_key:
            st.error("❌ 請上傳模板及配置 API Key")
        else:
            try:
                # 使用 BytesIO 避免路徑報錯
                wb = openpyxl.load_workbook(uploaded_tpl, keep_vba=True)
                sheet = wb.active
                h = {str(c.value).strip().lower().replace(" ", ""): c.column for r in sheet.iter_rows(max_row=3) for c in r if c.value}
                bp_cols = [c.column for r in sheet.iter_rows(max_row=3) for c in r if "keyproductfeatures" in str(c.value).lower().replace(" ", "")]
                client = OpenAI(api_key=api_key)
                curr_row = 5 # 子體從第 5 行開始

                for idx, item in enumerate(sku_inputs):
                    if not item["pfx"] or not item["img"]: continue
                    # 圖像指針復位
                    item["img"].seek(0)
                    b64 = base64.b64encode(item["img"].read()).decode('utf-8')
                    res = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[{"role":"user","content":[{"type":"text","text":"Analyze art JSON: {'title':'','elements':'','color':'','bp':['','','','','']}"},{"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{b64}"}}]}],
                        response_format={"type":"json_object"}
                    )
                    ai = json.loads(res.choices[0].message.content)
                    
                    # 規則鎖定：Parent SKU 命名
                    p_sku = f"{item['pfx']}-{n1}-{n3}"
                    
                    # 1 父 + 3 子 邏輯
                    rows_logic = [
                        {"type": "P", "sku": p_sku, "sz": "", "pr": ""},
                        {"type": "C", "sku": f"{item['pfx']}-{n1}", "sz": s1, "pr": p1},
                        {"type": "C", "sku": f"{item['pfx']}-{n2}", "sz": s2, "pr": p2},
                        {"type": "C", "sku": f"{item['pfx']}-{n3}", "sz": s3, "pr": p3}
                    ]

                    for r_data in rows_logic:
                        # 鎖定：第一行數據永遠寫在 Row 4
                        target_row = 4 if r_data["type"] == "P" else curr_row
                        def fill(k, v):
                            target = [col for name, col in h.items() if k.lower().replace(" ", "") in name]
                            if target:
                                sheet.cell(row=target_row, column=target[0], value=strict_clean(v))

                        fill("sellersku", r_data["sku"])
                        fill("parentsku", p_sku)
                        
                        if r_data["type"] == "C":
                            color_v = f"{ai.get('color','')} {ai.get('elements','')}"
                            fill("color", color_v)
                            fill("colormap", color_v)
                            fill("size", r_data["sz"])
                            fill("sizemap", r_data["sz"])
                            fill("standardprice", r_data["pr"])
                        
                        fill("productname", f"{brand} {ai.get('title','')} {ai.get('elements','')}"[:199])
                        fill("generickeyword", format_kw_strict(f"{ai.get('elements','')} {user_kw}"))
                        # 五點描述對位
                        for i in range(5):
                            fill(f"keyproductfeatures{i+1}", ai['bp'][i] if i < len(ai['bp']) else "")

                        if r_data["type"] == "C": curr_row += 1

                st.success("✅ AI 填充完成！")
                out = io.BytesIO()
                wb.save(out)
                st.download_button("💾 下載生成文件", out.getvalue(), "Amazon_Listing.xlsm")
            except Exception as e:
                st.error(f"❌ 錯誤: {e}")

# ==========================================
# 模式二：站點搬運 (獨立模塊，解決圖 ba77 報錯)
# ==========================================
elif mode == "站點搬運 (US ➔ UK)":
    st.header("🌍 跨站點數據自動搬運 (US ➔ UK)")
    st.info("將 US 已填表格搬運至 UK 空白模板。")

    col_us, col_uk = st.columns(2)
    with col_us:
        us_file = st.file_uploader("📂 1. 上傳已填寫的美國站表格 (US)", type=['xlsx', 'xlsm'], key="us_v145")
    with col_uk:
        uk_tpl = st.file_uploader("📂 2. 上傳空白的英國站模板 (UK)", type=['xlsx', 'xlsm'], key="uk_v145")

    if st.button("🚀 執行站點搬運", type="primary", key="move_btn"):
        if not us_file or not uk_tpl:
            st.error("❌ 請同時上傳兩個站點的文件")
        else:
            try:
                # 數據讀取
                us_wb = openpyxl.load_workbook(us_file, data_only=True)
                us_sheet = us_wb.active
                uk_wb = openpyxl.load_workbook(uk_tpl, keep_vba=True)
                uk_sheet = uk_wb.active

                # 表頭映射
                us_h = {str(c.value).strip().lower().replace(" ", ""): c.column for c in us_sheet[3] if c.value}
                uk_h = {str(c.value).strip().lower().replace(" ", ""): c.column for c in uk_sheet[3] if c.value}

                mapping = {
                    "sellersku": "sellersku", "parentsku": "parentsku",
                    "productname": "itemname", "brandname": "brandname",
                    "productdescription": "productdescription",
                    "generickeyword": "searchterms", "color": "colour",
                    "colormap": "colourmap", "size": "size", "sizemap": "sizemap",
                    "standardprice": "standardprice", "mainimageurl": "mainimageurl"
                }

                for r_idx in range(4, us_sheet.max_row + 1):
                    sku_check = us_sheet.cell(row=r_idx, column=us_h.get('sellersku', 1)).value
                    if not sku_check: continue

                    for us_k, uk_k in mapping.items():
                        if us_k in us_h and uk_k in uk_h:
                            val = us_sheet.cell(row=r_idx, column=us_h[us_k]).value
                            uk_sheet.cell(row=r_idx, column=uk_h[uk_k], value=strict_clean(val))
                    
                    # 搬運五點
                    for i in range(1, 6):
                        u_col = us_h.get(f"keyproductfeatures{i}") or us_h.get(f"bulletpoint{i}")
                        k_col = uk_h.get(f"bulletpoint{i}") or uk_h.get(f"keyproductfeatures{i}")
                        if u_col and k_col:
                            uk_sheet.cell(row=r_idx, column=k_col, value=us_sheet.cell(row=r_idx, column=u_col).value)

                st.success("✅ 站點數據搬運成功！")
                out_uk = io.BytesIO()
                uk_wb.save(out_uk)
                st.download_button("💾 下載轉換後文件", out_uk.getvalue(), "UK_Transfer.xlsm")
            except Exception as e:
                st.error(f"❌ 搬运失败: {e}")
