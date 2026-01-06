import streamlit as st
import pandas as pd
import io, base64, json, re, openpyxl, os
from openai import OpenAI
from openpyxl.styles import Font

# --- 1. 核心工具：格式與數據清洗 ---
def clean_text(text):
    """徹底清除 JSON 符號、AI 佔位詞及雜質"""
    if pd.isna(text) or text == "": return ""
    text = re.sub(r"[\[\]'\"']", "", str(text))
    # 過濾常見 AI 廢話
    blacklist = {'word1', 'word2', 'fake', 'placeholder', 'detailed', 'rich'}
    words = str(text).split()
    return " ".join([w for w in words if w.lower() not in blacklist]).strip()

def format_kw_strict(raw_text):
    """關鍵詞規則：僅空格分隔，限長 245 字符"""
    if not raw_text: return ""
    clean = re.sub(r'[^a-zA-Z0-9\s]', ' ', str(raw_text).lower())
    seen, res = set(), []
    for w in clean.split():
        if w not in seen and len(w) > 1:
            res.append(w)
            seen.add(w)
    return " ".join(res)[:245]

# --- 2. 頁面配置與環境變量讀取 ---
st.set_page_config(page_title="亞馬遜全能工具 V23.5", layout="wide")

# 優先讀取 Codespaces 終端注入的 Key
api_key = os.getenv("OPENAI_API_KEY") or st.secrets.get("OPENAI_API_KEY") or ""

# --- 3. 功能導航 ---
mode = st.sidebar.radio("功能導航", ["批量上架 (圖片分析)", "站點搬運 (US ➔ UK)"])

# ==========================================
# 模式一：批量上架 (鎖定 Row 4 為父體)
# ==========================================
if mode == "批量上架 (圖片分析)":
    st.header("🎨 AI 視覺分析上架模塊")
    with st.sidebar:
        st.subheader("⚙️ 規格鎖定")
        brand = st.text_input("品牌名稱", "AMAZING WALL")
        s1, p1 = st.text_input("尺寸 1", "16x24\""), st.text_input("價格 1", "12.99")
        s2, p2 = st.text_input("尺寸 2", "24x36\""), st.text_input("價格 2", "16.99")
        s3, p3 = st.text_input("尺寸 3", "32x48\""), st.text_input("價格 3", "19.99")

    pfx = st.text_input("SKU 前綴 (例: LMX-SDS)")
    img_file = st.file_uploader("上傳分析圖")
    tpl_file = st.file_uploader("上傳 Amazon 1.3MB 模板", key="tpl_us")

    if st.button("🚀 啟動 AI 填充", type="primary") and img_file and tpl_file and api_key:
        with st.spinner('正在分析並寫入 Template 子表...'):
            try:
                img_file.seek(0)
                b64 = base64.b64encode(img_file.read()).decode('utf-8')
                client = OpenAI(api_key=api_key)
                res = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role":"user","content":[{"type":"text","text":"Analyze art JSON: {'title':'','elements':'','color':'','bp':['','','','','']}"},{"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{b64}"}}]}],
                    response_format={"type":"json_object"}
                )
                ai = json.loads(res.choices[0].message.content)
                
                wb = openpyxl.load_workbook(tpl_file, keep_vba=True)
                # 鎖定 Template 子表
                sheet = wb['Template'] if 'Template' in wb.sheetnames else wb.active
                h = {str(c.value).strip().lower().replace(" ", ""): c.column for r in sheet.iter_rows(max_row=3) for c in r if c.value}
                
                p_sku = f"{pfx}-001-003"
                rows_data = [
                    {"type": "P", "sku": p_sku, "sz": "", "pr": ""},
                    {"type": "C", "sku": f"{pfx}-001", "sz": s1, "pr": p1},
                    {"type": "C", "sku": f"{pfx}-002", "sz": s2, "pr": p2},
                    {"type": "C", "sku": f"{pfx}-003", "sz": s3, "pr": p3}
                ]

                for i, r_info in enumerate(rows_data):
                    target_row = 4 + i
                    def fill(k, v):
                        col_indices = [idx for name, idx in h.items() if k.lower().replace(" ", "") in name]
                        if col_indices: sheet.cell(row=target_row, column=col_indices[0], value=clean_text(v))

                    fill("sellersku", r_info["sku"])
                    fill("parentsku", p_sku)
                    if r_info["type"] == "C":
                        cv = f"{ai.get('color','')} {ai.get('elements','')}"
                        fill("color", cv); fill("colormap", cv)
                        fill("size", r_info["sz"]); fill("sizemap", r_info["sz"])
                        fill("standardprice", r_info["pr"])
                    fill("productname", f"{brand} {ai.get('title','')} {ai.get('elements','')}"[:199])
                    for bi in range(5):
                        fill(f"keyproductfeatures{bi+1}", ai['bp'][bi] if bi < len(ai['bp']) else "")

                out = io.BytesIO()
                wb.save(out)
                st.success("✅ 美國站上架文件生成成功！")
                st.download_button("💾 下載結果", out.getvalue(), f"{pfx}_US.xlsm")
            except Exception as e: st.error(f"❌ 出錯: {e}")

# ==========================================
# 模式二：站點搬運 (精準子表對位版)
# ==========================================
elif mode == "站點搬運 (US ➔ UK)":
    st.header("🌍 跨站點精準數據搬運 (US ➔ UK)")
    st.info("💡 規則：系統會自動搜索 'Template' 子表，並將美國站數據映射至英國站對應列。")
    
    us_data = st.file_uploader("📂 1. 上傳已填好的 US 文件")
    uk_tpl = st.file_uploader("📂 2. 上傳空白 UK 模板")

    if st.button("🚀 執行精準搬運", type="primary") and us_data and uk_tpl:
        with st.spinner('正在掃描 US 數據並寫入 UK Template 子表...'):
            try:
                # 1. 讀取 US 的 Template 數據
                us_xl = pd.ExcelFile(us_data)
                us_sheet_name = 'Template' if 'Template' in us_xl.sheet_names else us_xl.sheet_names[0]
                us_df = pd.read_excel(us_data, sheet_name=us_sheet_name, header=2) 

                # 2. 讀取 UK 模板並定位 Template 表
                uk_wb = openpyxl.load_workbook(uk_tpl, keep_vba=True)
                uk_sheet = uk_wb['Template'] if 'Template' in uk_wb.sheetnames else uk_wb.active
                uk_h = {str(c.value).strip().lower().replace(" ", ""): c.column for c in uk_sheet[3] if c.value}

                # 3. 定義精準映射表 (解決標題、拼寫、關鍵詞差異)
                field_mapping = {
                    "productname": "itemname",           # 標題對位
                    "generickeywords": "searchterms",    # 關鍵詞對位
                    "color": "colour",                   # 英式拼寫
                    "colormap": "colourmap",
                    "standardprice": "standardprice",
                    "productdescription": "productdescription"
                }

                # 4. 循環搬運
                for col in us_df.columns:
                    src_clean = str(col).strip().lower().replace(" ", "")
                    # 映射處理
                    tgt_name = field_mapping.get(src_clean, src_clean)
                    
                    if tgt_name in uk_h:
                        col_idx = uk_h[tgt_name]
                        data_list = us_df[col].tolist()
                        for r_idx, val in enumerate(data_list, start=4):
                            uk_sheet.cell(row=r_idx, column=col_idx, value=clean_text(val))
                
                # 5. 五點描述特殊對位 (Bullet Points)
                for i in range(1, 6):
                    us_bp_key = f"keyproductfeatures{i}"
                    uk_bp_key = f"bulletpoint{i}" # 英國模板常用名
                    
                    # 檢查 UK 模板實際列名
                    target_bp_col = uk_h.get(uk_bp_key) or uk_h.get(us_bp_key)
                    
                    if target_bp_col:
                        # 尋找 US 表中匹配的列 (忽略大小寫)
                        us_col_match = [c for c in us_df.columns if us_bp_key in str(c).lower().replace(" ","")]
                        if us_col_match:
                            for r_idx, val in enumerate(us_df[us_col_match[0]].tolist(), start=4):
                                uk_sheet.cell(row=r_idx, column=target_bp_col, value=clean_text(val))

                out_uk = io.BytesIO()
                uk_wb.save(out_uk)
                st.success("✅ 搬運成功！數據已精準映射至 UK 站 Template 子表。")
                st.download_button("💾 下載英國站轉換文件", out_uk.getvalue(), "Amazon_UK_Final.xlsm")
                
            except Exception as e:
                st.error(f"❌ 搬運失敗: {str(e)}")
